"""
Endpoints DB-driven (RF-010): exponen planes y estudiantes desde la DB
y permiten correr el motor sin que el front mande JSON gigantes.
"""

from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from pydantic import ValidationError

from kairos.api.deps import get_db
from kairos.api.schemas.optimizer import (
    EscenarioReporte,
    RequestReporteComparativo,
    ResponsePrescripcion,
    ResponseReporteComparativo,
)
from kairos.core.optimizer import KairosOptimizer
from kairos.db.models import PlanORM
from kairos.db.repository import (
    AulaRepository,
    DocenteRepository,
    EstudianteRepository,
    HistoricoRepository,
    PlanRepository,
    RecursoRepository,
)
from kairos.schemas.data_models import (
    Aula,
    ConfiguracionKairos,
    Docente,
    EstudianteTrayectoria,
    PlanEstudio,
)

router = APIRouter()


def _construir_optimizer(
    db: Session,
    codigo_plan: str,
    config: Optional[ConfiguracionKairos],
) -> KairosOptimizer:
    """
    Levanta plan + estudiantes + recursos operativos (comisiones, aulas,
    docentes, histórico) de la DB y arma un KairosOptimizer listo para correr.

    Lanza HTTPException 404 si el plan no existe y 400 si el plan tiene ciclos.
    Centraliza la carga para que /process, /export y /reportes usen lo mismo.
    """
    plan = PlanRepository(db).get(codigo_plan)
    if plan is None:
        raise HTTPException(status_code=404, detail=f"Plan {codigo_plan} no existe")

    optimizer = KairosOptimizer(plan, config)
    for est in EstudianteRepository(db).list_by_plan(codigo_plan):
        optimizer.agregar_estudiante(est)

    recursos = RecursoRepository(db).list_all()
    if recursos:
        optimizer.agregar_recursos(recursos)

    aulas = AulaRepository(db).list_all()
    if aulas:
        optimizer.agregar_aulas(aulas)

    docentes = DocenteRepository(db).list_all()
    if docentes:
        optimizer.agregar_docentes(docentes)

    historico = HistoricoRepository(db).list_all()
    if historico:
        optimizer.agregar_historico(historico)

    bardo_plan = optimizer.validar_caminos()
    if any("ciclos" in p.lower() for p in bardo_plan):
        raise HTTPException(
            status_code=400,
            detail=f"El plan de estudio tiene ciclos: {bardo_plan}",
        )

    return optimizer


class PlanResumen(BaseModel):
    """Datos basicos de un plan, sin materias."""

    codigo_plan: str
    nombre_carrera: str
    facultad: Optional[str] = None
    ano_vigencia: int
    duracion_anos: int
    total_creditos: Optional[int] = None
    cantidad_materias: int


class ProcessFromDbRequest(BaseModel):
    config: Optional[ConfiguracionKairos] = None


class IngestaEstudiantesResponse(BaseModel):
    persistidos: int
    rechazados: int
    errores: List[str] = []


class IngestaResponse(BaseModel):
    """Respuesta genérica de una ingesta batch (aulas, docentes, etc)."""

    persistidos: int
    rechazados: int
    errores: List[str] = []


@router.post("/planes", response_model=PlanResumen, status_code=201)
def ingestar_plan(plan: PlanEstudio, db: Session = Depends(get_db)) -> PlanResumen:
    """
    Ingesta un plan de estudio (idempotente: pisa el existente si el codigo
    ya esta en la DB).
    """
    repo = PlanRepository(db)
    repo.upsert(plan)
    return PlanResumen(
        codigo_plan=plan.codigo_plan,
        nombre_carrera=plan.nombre_carrera,
        facultad=plan.facultad,
        ano_vigencia=plan.ano_vigencia,
        duracion_anos=plan.duracion_anos,
        total_creditos=plan.total_creditos,
        cantidad_materias=len(plan.materias),
    )


@router.post("/estudiantes", response_model=IngestaEstudiantesResponse, status_code=201)
def ingestar_estudiantes(
    estudiantes: List[dict], db: Session = Depends(get_db)
) -> IngestaEstudiantesResponse:
    """
    Ingesta una lista de estudiantes (idempotente por estudiante_id).
    Items invalidos se rechazan individualmente sin abortar el batch.
    """
    repo = EstudianteRepository(db)
    persistidos = 0
    errores: List[str] = []
    for item in estudiantes:
        try:
            est = EstudianteTrayectoria(**item)
        except ValidationError as e:
            errores.append(f"{item.get('estudiante_id', '?')}: {e.error_count()} errores")
            continue
        repo.upsert(est)
        persistidos += 1
    return IngestaEstudiantesResponse(
        persistidos=persistidos,
        rechazados=len(errores),
        errores=errores,
    )


@router.post("/aulas", response_model=IngestaResponse, status_code=201)
def ingestar_aulas(aulas: List[dict], db: Session = Depends(get_db)) -> IngestaResponse:
    """
    Ingesta una lista de aulas (idempotente por aula_id).
    Items invalidos se rechazan individualmente sin abortar el batch.
    """
    repo = AulaRepository(db)
    persistidos = 0
    errores: List[str] = []
    for item in aulas:
        try:
            aula = Aula(**item)
        except ValidationError as e:
            errores.append(f"{item.get('aula_id', '?')}: {e.error_count()} errores")
            continue
        repo.upsert(aula)
        persistidos += 1
    return IngestaResponse(persistidos=persistidos, rechazados=len(errores), errores=errores)


@router.post("/docentes", response_model=IngestaResponse, status_code=201)
def ingestar_docentes(docentes: List[dict], db: Session = Depends(get_db)) -> IngestaResponse:
    """
    Ingesta una lista de docentes (idempotente por docente_id).
    Items invalidos se rechazan individualmente sin abortar el batch.
    """
    repo = DocenteRepository(db)
    persistidos = 0
    errores: List[str] = []
    for item in docentes:
        try:
            docente = Docente(**item)
        except ValidationError as e:
            errores.append(f"{item.get('docente_id', '?')}: {e.error_count()} errores")
            continue
        repo.upsert(docente)
        persistidos += 1
    return IngestaResponse(persistidos=persistidos, rechazados=len(errores), errores=errores)


@router.get("/planes", response_model=List[PlanResumen])
def listar_planes(db: Session = Depends(get_db)) -> List[PlanResumen]:
    """Lista todos los planes guardados, con metadata basica."""
    rows = db.execute(select(PlanORM)).scalars().all()
    return [
        PlanResumen(
            codigo_plan=p.codigo_plan,
            nombre_carrera=p.nombre_carrera,
            facultad=p.facultad,
            ano_vigencia=p.ano_vigencia,
            duracion_anos=p.duracion_anos,
            total_creditos=p.total_creditos,
            cantidad_materias=len(p.materias),
        )
        for p in rows
    ]


@router.get("/planes/{codigo_plan}", response_model=PlanEstudio)
def obtener_plan(codigo_plan: str, db: Session = Depends(get_db)) -> PlanEstudio:
    """Devuelve un plan completo (con materias y correlativas)."""
    plan = PlanRepository(db).get(codigo_plan)
    if plan is None:
        raise HTTPException(status_code=404, detail=f"Plan {codigo_plan} no existe")
    return plan


@router.get(
    "/planes/{codigo_plan}/estudiantes",
    response_model=List[EstudianteTrayectoria],
)
def listar_estudiantes(
    codigo_plan: str, db: Session = Depends(get_db)
) -> List[EstudianteTrayectoria]:
    """Devuelve los estudiantes de un plan con todos sus registros de trayectoria."""
    if db.get(PlanORM, codigo_plan) is None:
        raise HTTPException(status_code=404, detail=f"Plan {codigo_plan} no existe")
    return EstudianteRepository(db).list_by_plan(codigo_plan)


@router.post(
    "/planes/{codigo_plan}/process",
    response_model=ResponsePrescripcion,
)
def procesar_desde_db(
    codigo_plan: str,
    request: ProcessFromDbRequest,
    db: Session = Depends(get_db),
) -> ResponsePrescripcion:
    """
    Levanta el plan y los estudiantes de la DB y corre el motor.
    Equivalente a POST /process pero sin que el front mande el JSON entero.
    """
    optimizer = _construir_optimizer(db, codigo_plan, request.config)

    prescripciones = optimizer.prescribir_aperturas()
    demanda = optimizer.analizar_demanda()
    cuellos = optimizer.detectar_cuellos_de_botella()
    resumen = optimizer.reporte_prescriptivo()
    metricas = optimizer.metricas_operativas(prescripciones)

    config_efectiva = request.config or ConfiguracionKairos()
    return ResponsePrescripcion(
        carrera=optimizer.plan.nombre_carrera,
        prescripciones=prescripciones,
        cuellos_botella=cuellos,
        demanda_total=sum(demanda.values()),
        materias_con_demanda=len(demanda),
        resumen=resumen,
        metricas_operativas=metricas,
        config_usada={
            "weight_tasa_graduacion": config_efectiva.weight_tasa_graduacion,
            "weight_eficiencia_operativa": config_efectiva.weight_eficiencia_operativa,
            "min_tasa_ocupacion": config_efectiva.min_tasa_ocupacion,
            "max_cupos_por_comision": config_efectiva.max_cupos_por_comision,
            "max_comisiones_a_abrir": config_efectiva.max_comisiones_a_abrir,
        },
    )


def _config_a_dict(config: ConfiguracionKairos) -> dict:
    """Resumen serializable de una config para devolver en los reportes."""
    return {
        "weight_tasa_graduacion": config.weight_tasa_graduacion,
        "weight_eficiencia_operativa": config.weight_eficiencia_operativa,
        "min_tasa_ocupacion": config.min_tasa_ocupacion,
        "max_cupos_por_comision": config.max_cupos_por_comision,
        "max_comisiones_a_abrir": config.max_comisiones_a_abrir,
        "respetar_capacidad_aulas": config.respetar_capacidad_aulas,
        "respetar_disponibilidad_docentes": config.respetar_disponibilidad_docentes,
        "usar_historico_docentes": config.usar_historico_docentes,
    }


@router.post(
    "/planes/{codigo_plan}/reportes/comparativo",
    response_model=ResponseReporteComparativo,
)
def reporte_comparativo(
    codigo_plan: str,
    request: RequestReporteComparativo,
    db: Session = Depends(get_db),
) -> ResponseReporteComparativo:
    """
    Corre el motor con 2-3 configuraciones distintas y devuelve métricas
    comparables (ingresos, completitud de aulas, alocación de docentes,
    demanda satisfecha, cobertura de cuellos de botella) para graficarlas
    lado a lado en la reportería.
    """
    if not (1 <= len(request.configuraciones) <= 3):
        raise HTTPException(
            status_code=400,
            detail="Mandá entre 1 y 3 configuraciones para comparar.",
        )

    carrera = ""
    escenarios = []
    for i, item in enumerate(request.configuraciones):
        config = item.config or ConfiguracionKairos()
        # Cada escenario reconstruye su optimizer: el motor consume recursos
        # (aulas/docentes) durante la corrida, así que no se puede reusar.
        optimizer = _construir_optimizer(db, codigo_plan, config)
        carrera = optimizer.plan.nombre_carrera

        prescripciones = optimizer.prescribir_aperturas()
        metricas = optimizer.metricas_operativas(prescripciones)

        nombre = item.nombre or f"Config {i + 1}"
        escenarios.append(
            EscenarioReporte(
                nombre=nombre,
                config_usada=_config_a_dict(config),
                metricas=metricas,
            )
        )

    return ResponseReporteComparativo(carrera=carrera, escenarios=escenarios)


@router.post("/planes/{codigo_plan}/export/excel")
def exportar_excel(
    codigo_plan: str,
    request: ProcessFromDbRequest,
    db: Session = Depends(get_db),
):
    """
    Corre la optimización y exporta la propuesta definitiva de comisiones a un archivo Excel
    con detalles de aulas, profesores y marcado de comisiones con bajo cupo (riesgo).
    """
    optimizer = _construir_optimizer(db, codigo_plan, request.config)
    plan = optimizer.plan

    prescripciones = optimizer.prescribir_aperturas()

    # Filtrar solo comisiones que se recomienda ABRIR
    comisiones_abrir = [
        p for p in prescripciones.values() if p["decision"] == "ABRIR"
    ]

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Propuesta Oferta Académica"

    # Estilos de diseño premium
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="595959")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    font_risk = Font(name="Segoe UI", size=10, bold=True, color="C00000")

    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Azul oscuro
    fill_risk = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Naranja/Amarillo muy claro
    fill_zebra = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid") # Azul grisáceo levísimo
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Fila 1: Vacía
    ws.append([])

    # Fila 2: Título principal
    ws.cell(row=2, column=2, value="KAIROS — Propuesta Definitiva de Oferta Académica").font = font_title
    ws.row_dimensions[2].height = 25

    # Fila 3: Información de Carrera
    ws.cell(row=3, column=2, value=f"Carrera: {plan.nombre_carrera} (Plan {plan.codigo_plan})").font = font_subtitle
    ws.row_dimensions[3].height = 18

    # Fila 4: Configuración del Motor Usada
    config_efectiva = request.config or ConfiguracionKairos()
    config_desc = (
        f"Configuración activa: Peso Cascada: {int(config_efectiva.weight_tasa_graduacion * 100)}% | "
        f"Peso Rentabilidad: {int(config_efectiva.weight_eficiencia_operativa * 100)}% | "
        f"Score Mínimo: {config_efectiva.min_tasa_ocupacion * 10:.1f} | "
        f"Tope comisiones: {config_efectiva.max_comisiones_a_abrir or 'Sin límite'}"
    )
    ws.cell(row=4, column=2, value=config_desc).font = font_subtitle
    ws.row_dimensions[4].height = 18

    # Fila 5: Vacía
    ws.append([])

    # Encabezados de tabla (Fila 6)
    headers = [
        "Materia Código",
        "Nombre Materia",
        "Turno",
        "Aula Asignada",
        "Profesor Requerido",
        "Demanda Proyectada",
        "Score Prescriptivo",
        "Estado / Alerta"
    ]

    header_row = 6
    ws.row_dimensions[header_row].height = 28
    for col_idx, h_text in enumerate(headers, start=2):
        cell = ws.cell(row=header_row, column=col_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # Agregar filas de datos (a partir de la fila 7)
    row_idx = 7
    turno_map = {"manana": "Mañana", "tarde": "Tarde", "noche": "Noche"}

    for item in comisiones_abrir:
        estado_text = "⚠️ Riesgo de baja inscripción" if item["bajo_cupo"] else "Normal"
        
        c_cod = ws.cell(row=row_idx, column=2, value=item["codigo"])
        c_cod.alignment = align_center
        
        c_nom = ws.cell(row=row_idx, column=3, value=item["nombre"])
        c_nom.alignment = align_left
        
        c_tur = ws.cell(row=row_idx, column=4, value=turno_map.get(item["turno"], item["turno"]))
        c_tur.alignment = align_center
        
        c_aul = ws.cell(row=row_idx, column=5, value=item["aula"])
        c_aul.alignment = align_center
        
        c_doc = ws.cell(row=row_idx, column=6, value=item["docente"])
        c_doc.alignment = align_left
        
        c_dem = ws.cell(row=row_idx, column=7, value=item["demanda"])
        c_dem.alignment = align_right
        
        c_sco = ws.cell(row=row_idx, column=8, value=item["score"])
        c_sco.alignment = align_right
        
        c_est = ws.cell(row=row_idx, column=9, value=estado_text)
        c_est.alignment = align_center

        # Aplicar estilos y formatos por celda
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        row_fill = fill_risk if item["bajo_cupo"] else (fill_zebra if is_even else fill_white)

        for col in range(2, 10):
            c = ws.cell(row=row_idx, column=col)
            c.font = font_data
            c.fill = row_fill
            c.border = border_thin
            if item["bajo_cupo"]:
                if col in (7, 9): # Destacar en rojo la demanda y el estado
                    c.font = font_risk

        row_idx += 1

    # Auto-ajustar el ancho de las columnas para que no se corte el texto
    for col in range(2, 10):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(6, row_idx):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    # Ajustes finos de columnas específicas
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['C'].width = 32 # Nombre materia
    ws.column_dimensions['F'].width = 24 # Profesor
    ws.column_dimensions['I'].width = 25 # Estado/Alerta

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"propuesta_oferta_{codigo_plan}.xlsx"
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
